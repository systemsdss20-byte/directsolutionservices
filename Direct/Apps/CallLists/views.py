import datetime
import json
from django.conf import settings
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..Procedure.forms import ExamForm
from .forms import CommentForm
from ..helpers.message import MessageResponse

from ..Consortium.models import Comments, Detail_RandomList


# Create your views here.
class RandomListCallView(LoginRequiredMixin, View):
    login_url = "Procedure:login"
    def get(self, request, *args, **kwargs):
        context = {}
        try:
            random_list = Detail_RandomList.objects.filter(random_list__show=True)
            table = True if kwargs.get("table") == "true" else False
            export = True if request.GET.get("export") == "true" else False
            status = request.GET.get("status") if request.GET.get("status") else None
            if status:
                if status == "uncontacted":
                    random_list = random_list.filter(Q(status='uncontacted') | Q(status__isnull=True))
                else:
                    random_list = random_list.filter(status=status)
                
            else:
                random_list = random_list.exclude(status='negative')
                   
            if table :
                return render(request, 'Call_List/RandomList/table.html', {'group_customers': self.GroupCustomers(random_list), 'total_records': random_list.count()})
            if export:
                result = self.export_to_excel(random_list=random_list)
                return HttpResponse(result, content_type='application/json', status=200)
            
            context = {'group_customers': self.GroupCustomers(random_list), 'total_records': random_list.count()}
        except Exception as e:
            print("RandomListCallView",e)
        return render(request, 'Call_List/RandomList/random-list.html', context)
    
        
    def export_to_excel(self, random_list):
        filename = 'Random_List_Export_{0}.xlsx'.format(datetime.datetime.today().strftime('%m%d%Y_%H%M%S'))
        filepath = '{0}/{1}'.format(settings.FILES_PDF, filename)
        try:
            book = Workbook()
            sheet = book.active
            sheet.title = 'Random List Export'

            # 1. Definir los Estilos de Formato
            # Estilo de Fuente: Negrita y Color Blanco (opcional)
            header_font = Font(bold=True, color="FFFFFF") 
            # Estilo de Relleno: Color de fondo (por ejemplo, Azul Marino)
            header_fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid') 
            # Estilo de Alineación: Centrado (opcional)
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Add Header
            sheet.append(('CUSTOMER NAME', 'CUSTOMER CONTACT', 'DRIVER NAME', 'DRIVER PHONE', 'TEST SUBSTANCES', 'TEST ALCOHOL', 'STATUS'))

            # 2. Iterar y Aplicar Formato a la Primera Fila (Fila 1)
            # El método .row_dimensions[1] nos permite acceder a la Fila 1.
            for cell in sheet[1]: # sheet[1] obtiene todas las celdas de la Fila 1
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
            # Opcional: Ajustar el ancho de las columnas
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter # Obtiene el nombre de la columna (A, B, C, ...)
                for cell in col:
                    try: # Maneja valores None o vacíos
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width


            # Add data
            for item in random_list:
                sheet.append(
                    (
                        item.customer.cusname, 
                        item.customer.contact1, 
                        item.driver.nombre, 
                        item.driver.phone, 
                        'Si' if item.test_substances else 'No', 
                        'Si' if item.test_alcohol else 'No', 
                        item.status
                    )
                )
            book.save(filepath)
            return JsonResponse({'filename': filename})
        except Exception as e:
            print("Export to Excel Error:", e)
            return JsonResponse({'message': 'Error generating Excel file', 'type': 'error'})

    def patch(self, request, *args, **kwargs):
        try:
            params = json.loads(request.body)
            last_color = color_status(params['last_status'])
            new_color = color_status(params['status'])
            detail_random_id = params['detail_random_id']
            Detail_RandomList.objects.filter(id=detail_random_id).update(status=params['status'], updated_at=datetime.datetime.now(), updated_by_id=request.user.id)
            return MessageResponse(data={'last_color': last_color, 'new_color': new_color, 'detail_random_id': detail_random_id},description='Updated successfully').success()
        except Exception as e:
            print("Random List Call View:", e)
            return MessageResponse(description='Internal Server Error').error()


    def GroupCustomers(self, random_list):
        try:
            customers = {}
            for selected in random_list:
                test_file = ""
                if selected.customer.idcustomer not in customers:
                    all_comments = Comments.objects.filter(detail_random_list=selected).order_by('-created_at')
                    comments_count = all_comments.count()
                    last_comment = all_comments.first()
                    last_note = last_comment.comment if last_comment else ""
                    last_note_at = last_comment.created_at.strftime('%d/%m %H:%M') if last_comment else ""

                    if selected.test_file is not None:
                        if hasattr(selected.test_file, 'path'):
                            test_file = selected.test_file.path.url
                            
                    customer = {
                        'customer_id': selected.customer.idcustomer,
                        'customer_name': selected.customer.cusname,
                        'customer_contact': selected.customer.contact1,
                        'customer_phone': selected.customer.mobile1,
                        'drivers': [{
                            'detail_id': selected.id,
                            'driver_id': selected.driver.iddriver,
                            'driver_name': selected.driver.nombre,
                            'driver_phone': selected.driver.phone,
                            'status': selected.status,
                            'color': color_status(selected.status),
                            'test_file': test_file,
                            'count_notes': comments_count,
                            'last_note': last_note,
                            'last_note_at': last_note_at,
                            'tests': [
                                {
                                    'substances': selected.test_substances,
                                    'alcohol': selected.test_alcohol
                                }
                            ]
                        }] if selected.driver else [],
                        'tests': {
                            'substances': selected.test_substances if selected.driver is None else "",
                            'alcohol': selected.test_alcohol if selected.driver is None else ""
                        },
                        'status': selected.status if selected.driver is None else "",
                        'color': color_status(selected.status) if selected.driver is None else "",
                        'detail_id': selected.id,
                        'test_file': test_file,
                        'count_notes': comments_count if selected.driver is None else "",
                        'last_note': last_note if selected.driver is None else "",
                        'last_note_at': last_note_at if selected.driver is None else ""

                    }
                    customers[selected.customer.idcustomer] = customer
                else:
                    if selected.driver:
                        all_comments = Comments.objects.filter(detail_random_list=selected).order_by('-created_at')
                        comments_count = all_comments.count()
                        last_comment = all_comments.first()
                        last_note = last_comment.comment if last_comment else ""
                        last_note_at = last_comment.created_at.strftime('%d/%m %H:%M') if last_comment else ""

                        test_file = ""
                        if selected.test_file is not None:
                            if hasattr(selected.test_file, 'path'):
                                test_file = selected.test_file.path.url
                        driver = {
                            'detail_id': selected.id,
                            'driver_id': selected.driver.iddriver,
                            'driver_name': selected.driver.nombre,
                            'driver_phone': selected.driver.phone,
                            'status': selected.status,
                            'color': color_status(selected.status),
                            'test_file': test_file,
                            'count_notes': comments_count,
                            'last_note': last_note,
                            'last_note_at': last_note_at,
                            'tests': [
                                {
                                    'substances': selected.test_substances,
                                    'alcohol': selected.test_alcohol
                                }
                            ]
                        }
                        customers[selected.customer.idcustomer]['drivers'].append(driver)
            return list(customers.values())
        except Exception as e:
            print("GroupCustomers",e)

def color_status(status):
    #Contacted - Pending
    if status == 'pending':
        return 'bg-red-lt'
    if status == 'uncontacted':
        return 'btn-warning'
    if status == 'negative':
        return 'bg-dark-lt'
    elif status == 'scheduled':
        return 'btn-info'
    elif status == 'paid':
        return 'btn-success'
    elif status == 'completed':
        return 'btn-primary'
    return 'btn-dark'

class CommentsRandomList(LoginRequiredMixin, View):
    login_url = "Procedure:login"

    def get(self, request, *args, **kwargs):
        if kwargs.get("detail_randomlist_id"):
            return self.get_comments(request, kwargs.get("detail_randomlist_id"))
        elif kwargs.get("id"):
            return self.get_comment(kwargs.get("id"))
        else:
            return MessageResponse(description="Data not found").error()

    def get_comment(self, id):
        try:
            comment = Comments.objects.only('comment').get(pk=id)
            return JsonResponse({'comment': comment.comment})  # Access the comment attribute
        except Comments.DoesNotExist:
            return MessageResponse(
                description="Comment not found").error()  # Handle the case where the comment doesn't exist
        except Exception as e:  # Catch other potential errors (database issues, etc.)
            return MessageResponse(description=f"An error occurred: {e}").error()

    def get_comments(self,request, id):
        detail_randomtest = Detail_RandomList.objects.get(id=id)
        comments = Comments.objects.filter(detail_random_list_id=id)
        context = {
            'customer': detail_randomtest.customer.cusname,
            'driver': detail_randomtest.driver.nombre if detail_randomtest.driver else "",
            'detail_random_list': id,
            'comments': comments,
            'created_by': request.user.id
        }
        return render(request, 'Call_List/RandomList/comments.html', context)

    def post(self, request, *args, **kwargs):
        try:
            form = CommentForm(request.POST)
            num_comments = Comments.objects.filter(detail_random_list=request.POST.get("detail_random_list")).values('detail_random_list').count()
            if form.is_valid():
                comment = form.save()
                num_comments = num_comments + 1
                data = {
                    'comment': comment.comment, 'created_at': comment.created_at.strftime("%b %d, %Y, %H:%M:%S"),
                    'updated_at': comment.updated_at,
                    'user': comment.created_by.fullname, 'avatar': str(comment.created_by.avatar),
                    'detail_random_list_id': comment.detail_random_list.id,
                    'total_comments': num_comments
                }
                return MessageResponse(data=data, description='Comment added successfully').success()
            else:
                errors = form.errors.as_json(escape_html=False)
                result = JsonResponse({'errors': errors})
                return HttpResponse(result, content_type='application/json', status=500)
        except Exception as e:
            print(e)
            return MessageResponse(description='Internal Server Error').error()

    def patch(self, request, *args, **kwargs):
        try:
            params = json.loads(request.body)
            comment = Comments.objects.get(pk=params['id'])
            comment.comment = params['comment']
            comment.updated_at = datetime.datetime.now()
            comment.save()
            # comment = Comments.objects.filter(id=params['id']).update(comment=params['comment'], updated_at=datetime.datetime.now())
            data = {
                "comment": comment.comment,
                "date_update": comment.updated_at.strftime('%b %d, %Y %H:%M')
            }
            return MessageResponse(data=data, description="Comment updated successfully").success()
        except Exception as e:
            print(e)
            return MessageResponse(description="Internal Server Error").error()

class FileUploadRandomTest(LoginRequiredMixin, View):
    login_url = "Procedure:login"
    template = 'CallList/RandomList/file-test.html'
    
    def get(self, request, *args, **kwargs):
        detail_randomlist_id = kwargs["detail_randomlist_id"]
        detail = Detail_RandomList.objects.only('driver', 'customer').get(id=detail_randomlist_id)
        form = ExamForm()
        context = {'form': form, 'driver': detail.driver, 'detail_randomlist_id': detail_randomlist_id}
        return render(request, "Call_List/RandomList/file-test.html", context)
    
    def post(self, request, *args, **kwargs):
        form = ExamForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                detail_randomlist_id = kwargs["detail_randomlist_id"]
                detail = Detail_RandomList.objects.get(id=detail_randomlist_id)
                exam = form.save(commit=False)
                exam.filename = request.FILES["path"].name
                exam.path = request.FILES["path"]
                exam.save()
                # Asignar el examen guardado al campo test_file
                detail.test_file = exam
                detail.updated_by_id = request.user.id
                detail.updated_at = datetime.datetime.now()
                detail.status = "completed"
                detail.save()
                return MessageResponse(data={"detail_randomlist_id": detail_randomlist_id, "path": exam.path.url},description="File uploaded successfully").success()
            except detail.DoesNotExist:
                return MessageResponse(description="Data not found").error()
            except detail.MultipleObjectsReturned:
                return MessageResponse(description="Multiple objects returned").error()
            except Exception as e:
                print(e)
                return MessageResponse(description="Internal Server Error").error()     
        else:
            return MessageResponse(description="Check the form").error()