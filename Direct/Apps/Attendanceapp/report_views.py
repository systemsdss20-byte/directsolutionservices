import datetime
import os

from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Alignment, Side
from openpyxl.utils import get_column_letter
from .models import Attendance, Employee


class Attendance_Report(View):

    def post(self, request, *args, **kwargs):
        date_from = datetime.datetime.strptime(request.POST.get('date_from'), '%m/%d/%Y') if request.POST.get(
            'date_from') else ''
        date_to = datetime.datetime.strptime(request.POST.get('date_to'), '%m/%d/%Y') if request.POST.get(
            'date_to') else ''
        if date_from and date_to:
            attendance_filtered = Attendance.objects.filter(date__range=[date_from, date_to]).order_by(
                'employee__surnames')
            employees = attendance_filtered.values('employee_id').distinct() if request.POST.get(
                'employee') == '--Select--' else [{'employee_id': request.POST.get('employee')}]
            request = self.excel_export(self,employees=employees, attendance_control=attendance_filtered)
            return HttpResponse(request, content_type='application/json', status=200)
        else:
            request = JsonResponse({'message': 'No filter date applied', 'type': 'error'})
            return HttpResponse(request, content_type='application/json', status=500)

    @staticmethod
    def excel_export(self, employees, attendance_control):
        first_lap = True
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        # Create Excel book
        file_name = 'Attendance_{0}.xlsx'.format(datetime.datetime.today().strftime('%m%d%Y'))
        file_path = '{0}/{1}'.format(settings.FILES_PDF, file_name)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
            book = Workbook()
            # Create Sheets
            for employee in employees:
                employee_attendance = attendance_control.filter(employee=employee['employee_id'])
                employee = Employee.objects.only('names', 'surnames').get(id=employee['employee_id'])
                if first_lap:
                    sheet = book.active
                    sheet.title = '{0} {1}'.format(employee.surnames.upper(), employee.names.upper())
                else:
                    sheet = book.create_sheet('{0} {1}'.format(employee.surnames.upper(), employee.names.upper()))
                # Add Header
                sheet.append(('DATE', 'CLOCK IN', 'LUNCH IN', 'LUNCH OUT', 'CLOCK OUT', 'TOTAL HOURS'))
                # Convert to timedelta to add time
                # total_hours = datetime.timedelta(hours=0, minutes=0, seconds=0) + datetime.timedelta(hours=employee_hour.hour, minutes=employee_hour.minute, seconds=employee_hour.second)
                # Add data
                for employee_assistance in employee_attendance:
                    sheet.append((
                        employee_assistance.date, employee_assistance.clock_in_at, employee_assistance.lunch_in_at,
                        employee_assistance.lunch_out_at, employee_assistance.clock_out_at,
                        employee_assistance.hours))
                # Apply styles on cells
                footer_font = Font(bold=True, name='Arial')
                for rows in sheet.iter_rows(min_row=1, max_row=len(employee_attendance) + 1, min_col=1, max_col=6):
                    for cell in rows:
                        if cell.row == 1:
                            cell.style = "Accent1"
                            cell.font = Font(bold=True, name='Arial', color='00FFFFFF')
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            cell.font = Font(name='Arial')
                            if cell.column == 1:
                                sheet.cell(cell.row, cell.column).number_format = 'mm/dd/yyyy'
                            if cell.column == 6:
                                sheet.cell(cell.row, cell.column).number_format = '[h]:mm:ss'
                        cell.border = thin_border
                # Add sum of column F
                sheet.cell(len(employee_attendance) + 2, 5, 'TOTAL').font = footer_font
                sheet.cell(len(employee_attendance) + 2, 6, "=SUM(F1:F%s)" % (len(employee_attendance) + 1)).font = footer_font
                sheet.cell(len(employee_attendance) + 2, 6).number_format = '[h]:mm:ss'
                sheet.column_dimensions[get_column_letter(6)].auto_size = True
                sheet.column_dimensions[get_column_letter(1)].auto_size = True
                first_lap = False
            book.save(file_path)
            return JsonResponse({'filename': file_name})
        except Exception as e:
            print(e)
            return JsonResponse({'message': 'Internal server error while creating file', 'type': 'error'})
