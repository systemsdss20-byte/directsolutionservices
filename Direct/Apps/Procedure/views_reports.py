import os
from datetime import date, datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.db.models import Q, Sum, F
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from openpyxl import Workbook

from .models import Projects, Customers, Invoice_paid, Invoices, NotesProjects, Services, User
from .views import receive_summary, summary_projects
from django.views.generic import View
from ... import settings


@login_required(login_url='Procedure:login')
def daily_chart(request):
    content_type = ContentType.objects.get_for_model(Invoice_paid)
    Permission.objects.get_or_create(
        name="Daily Chart",
        codename="daily_chart",
        content_type=content_type,
    )
    if request.method == 'GET':
        customers = Customers.objects.all().only('idcustomer', 'cusname')
        return render(request, 'Procedure/Accounting/daily_chart.html', {'customers': customers})
    if  request.method == 'POST':
        sldate = int(request.POST.get('sldate'))
        idcustomer = int(request.POST.get('customers'))
        if idcustomer == 0:
            main_paid = Invoice_paid.objects.select_related('idinvoice').select_related('idinvoice__idcustomer').values(
                'idpaid', 'paid', 'typepaid',
                'comment', 'idinvoice', 'idinvoice__paid_date', 'idinvoice__amount', 'idinvoice__deu',
                'idinvoice__idcustomer_id',
                'idinvoice__idcustomer__cusname').filter(idinvoice__deleted=0)
            #main_invoice = Invoices.objects.filter(deleted=0).filter(~Q(amount=F('deu'))).only('amount')
        else:
            main_paid = Invoice_paid.objects.select_related('idinvoice').select_related('idinvoice__idcustomer').values(
                'idpaid', 'paid', 'typepaid',
                'comment', 'idinvoice', 'idinvoice__paid_date', 'idinvoice__amount', 'idinvoice__deu',
                'idinvoice__idcustomer_id',
                'idinvoice__idcustomer__cusname').filter(idinvoice__idcustomer_id=idcustomer, idinvoice__deleted=0)
            #main_invoice = Invoices.objects.filter(deleted=0).only('amount')
        if sldate == 1:
            start_date = datetime.strptime(request.POST.get('date'), '%m/%d/%Y').date()
            end_date = datetime.strptime(request.POST.get('second_date'), '%m/%d/%Y').date()
            paids = main_paid.filter(datepaid__range=(start_date, end_date))
        if sldate == 2:
            filter_date = request.POST.get('date').split('/')
            year = filter_date[1]
            month = filter_date[0]
            paids = main_paid.filter(datepaid__year__gte=year, datepaid__month=month, datepaid__year__lte=year)
        if sldate == 3:
            filter_year = request.POST.get('date')
            paids = main_paid.filter(datepaid__year=filter_year)
        if sldate == 0:
            year = date.today().year
            paids = main_paid.filter(datepaid__year=year)
        data = list()
        for paid in paids:
            customer = '{}-{}'.format(paid['idinvoice__idcustomer_id'], paid['idinvoice__idcustomer__cusname'])
            invdate = paid['idinvoice__paid_date']
            action = '<a class="btn btn-sm btn-pill btn-success" href="/Procedure/pdfInvoice/{}"><li class="fa fa-eye"></li>&nbsp;View</a>'.format(
                paid["idinvoice"])
            data.append({'id': paid['idinvoice'], 'customer': customer, 'paid_date': invdate.strftime("%m/%d/%Y"),
                         'typepaid': paid['typepaid'], 'comment': paid['comment'],
                         'invamount': paid['idinvoice__amount'], 'paid': paid['paid'], 'invdeu': paid['idinvoice__deu'],
                         'action': action})
        num = paids.count()
        sum_cash = paids.filter(typepaid='Cash').aggregate(paid_s=Sum('paid'))
        sum_check = paids.filter(typepaid='Check').aggregate(paid_s=Sum('paid'))
        sum_credit_card = paids.filter(typepaid='Credit Card').aggregate(paid_s=Sum('paid'))
        sum_zelle = paids.filter(typepaid='Zelle').aggregate(paid_s=Sum('paid'))
        total_paid = paids.aggregate(paid_s=Sum('paid'))
        invoices = paids.values('idinvoice', 'idinvoice__amount').distinct()
        total_amount = invoices.aggregate(amount_s=Sum('idinvoice__amount'))
        total_due = paids.aggregate(due_s=Sum('idinvoice__deu'))
        return JsonResponse({'cash': sum_cash['paid_s'], 'checks': sum_check['paid_s'],
                             'credit_card': sum_credit_card['paid_s'], 'zelle': sum_zelle['paid_s'],
                             'total_paid': total_paid['paid_s'], 'total_amount': total_amount['amount_s'],
                             'total_due': total_due['due_s'], 'data': data, 'draw': 1, 'recordsTotal': num,
                             'recordsFiltered': num}, safe=False)


'''@login_required(login_url='Procedure:login')
def invoice_unpaid(request):
    if request.method == 'GET':
        return render(request, 'Procedure/Accounting/invoices_unpaid.html')
    if  request.method == 'POST':
        invoices = Invoices.objects.all().filter(status='unpaid', idcustomer__clientstatus='Active')
        data = list()
        for invoice in invoices:
            data.append({'id': invoice.idinvoice, 'customer': '%s - %s' % (invoice.idcustomer_id, invoice.cusname),
                         'invdate': invoice.invdate.strftime("%m/%d/%Y"), 'amount': invoice.amount, 'due': invoice.deu
                         })
        return JsonResponse({'data': data}, safe=False)
'''


class Invoice_Unpaid(View):
    template_name = 'Procedure/Accounting/invoices_unpaid.html'

    def get(self, request, *args, **kwargs):
        summary = summary_projects(request)
        return render(request, self.template_name, {'summary': summary})


@login_required(login_url='Procedure:login')
def projects_reports(request):
    content_type = ContentType.objects.get_for_model(Projects)
    Permission.objects.get_or_create(
        name="Projects Report",
        codename="projects_report",
        content_type=content_type,
    )
    if request.method == 'GET':
        date_now = datetime.now()
        last_seven_days = date_now - timedelta(days=7)
        customers = Customers.objects.filter(clientstatus='Active').only('idcustomer', 'cusname')
        summary = summary_projects(request)
        services = Services.objects.filter(is_active=True)
        users = User.objects.filter(is_active=True).exclude(username='admin').only('id', 'fullname')
        return render(request, 'Procedure/Reports/projects_reports.html', {
            'customers': customers, 'summary': summary, 'last_seven_days': last_seven_days.strftime('%m/%d/%Y'),
            'services': services, 'users': users
        })

    if  request.method == 'POST':
        idcustomer = int(request.POST.get('customer'))
        status = int(request.POST.get('status'))
        service_searched = int(request.POST.get('service'))
        last_user = int(request.POST.get('lastuser'))
        all_date = True if request.POST.get('all_date') == 'true' else False
        
        if idcustomer == 0:
            projects = Projects.objects.filter(deleted=0, idcustomer__clientstatus='Active')
        else:
            projects = Projects.objects.filter(deleted=0, idcustomer__clientstatus='Active', idcustomer=idcustomer)
        if last_user != 0:
            projects = projects.filter(iduserlast=last_user)
        if service_searched != 0:
            projects = projects.filter(service__idservice=service_searched)
        if status == 1:
            projects = projects.filter(Q(status='Opened') | Q(status='In Process'))
        if status == 2:
            projects = projects.filter(status='Opened')
        if status == 3:
            projects = projects.filter(status='In Process')
        if status == 4:
            projects = projects.filter(status='Closed')
        if request.POST.get('from_date') and all_date is False:
            from_date = datetime.strptime(request.POST.get('from_date'), '%m/%d/%Y').date()
            to_date = datetime.strptime(request.POST.get('to_date'), '%m/%d/%Y').date()
            projects = projects.filter(request__range=(from_date, to_date))

        data = list()
        for json in projects:
            num_notes = NotesProjects.objects.filter(project=json).count()
            data.append({
                'id': json.idproject,
                'customers': {'idcustomer': json.idcustomer_id, 'cusname': json.idcustomer.cusname},
                'services': json.service_name,
                'comments': json.comments,
                'representative': (json.iduser.first_name + " " + json.iduser.last_name),
                'invoice': json.idinvoicedet.idinvoice.status if json.idinvoicedet else 'No Invoice', 'num_notes': num_notes,
                'request': json.request.strftime('%m/%d/%Y'), 'status': json.status,
                'userlast': (json.iduserlast.first_name + " " + json.iduserlast.last_name).upper()
            })
        num = projects.count()
        return JsonResponse({'draw': 1, 'recordsTotal': num, 'recordsFiltered': num, 'data': data}, safe=False)


@login_required(login_url='Procedure:login')
def expirations(request):
    content_type = ContentType.objects.get_for_model(Customers)
    Permission.objects.get_or_create(
        name="Expirations Reports",
        codename="expirations_reports",
        content_type=content_type,
    )
    if request.user.is_authenticated:
        if request.method == 'GET':
            type_expirations = list([{'label': 'Local Tag', 'value': 'TAG'}, {'label': 'IRP', 'value': 'IRP'},
                                     {'label': 'IFTA & Quarter', 'value': 'IFTA'},
                                     {'label': 'Insurance', 'value': 'INS'},
                                     {'label': 'Anual Report', 'value': 'ANR'},
                                     {'label': 'New Mexico WDT', 'value': 'NM'},
                                     {'label': 'New York Hut', 'value': 'NY'}, {'label': 'Kentucky', 'value': 'KYU'},
                                     {'label': 'UCR', 'value': 'UCR'}, {'label': 'DOT Update', 'value': 'DOT'},
                                     {'label': 'Random', 'value': 'RANDOM'},
                                     {'label': 'Inner Bridge', 'value': 'INNER BRIDGE'},
                                     {'label': 'Overweight', 'value': 'OVERWEIGHT'},
                                     {'label': 'Oversize', 'value': 'OVERSIZE'},
                                     ])
            return render(request, 'Procedure/Reports/expirations.html', {'types': type_expirations})
        if  request.method == 'POST':
            type_search = str(request.POST.get('type'))
            by = str(request.POST.get('by'))
            from_date = datetime.strptime(request.POST.get('from_date'), '%m/%d/%Y').date()
            to_date = datetime.strptime(request.POST.get('to_date'), '%m/%d/%Y').date()
            data = list()
            if type_search == 'TAG':
                customers = Customers.objects.filter(clientstatus='Active').only('idcustomer', 'cusname', 'contact1',
                                                                                 'email', 'mobile1', 'floridaid',
                                                                                 'floridaexp')
                if by == 'all':
                    customers = customers.exclude(floridaexp='0001-01-01')
                if by == 'date':
                    customers = customers.filter(floridaexp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.floridaid,
                         'dateexp': customer.floridaexp.strftime('%m/%d/%Y')})
            if type_search == 'IRP':
                customers = Customers.objects.filter(clientstatus='Active').only('idcustomer', 'cusname', 'contact1',
                                                                                 'email', 'mobile1', 'irpid',
                                                                                 'irpexp').order_by('irpexp')
                if by == 'all':
                    customers = customers.exclude(irpexp='0001-01-01')
                if by == 'date':
                    customers = customers.filter(irpexp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.irpid,
                         'dateexp': customer.irpexp.strftime('%m/%d/%Y')})
            if type_search == 'IFTA':
                customers = Customers.objects.filter(clientstatus='Active').exclude(
                    Q(iftaid__exact='') | Q(iftaid__isnull=True)).only('idcustomer', 'cusname', 'contact1', 'email',
                                                                       'mobile1', 'iftaid', 'iftaexp').order_by(
                    'iftaexp')
                if by == 'date':
                    customers = customers.filter(iftaexp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.iftaid,
                         'dateexp': customer.iftaexp.strftime('%m/%d/%Y')})
            if type_search == 'INS':
                if by == 'all':
                    customers = Customers.objects.filter(clientstatus='Active').exclude(insuexpire='0001-01-01').only(
                        'idcustomer', 'cusname', 'contact1', 'email', 'mobile1', 'insuexpire')
                if by == 'date':
                    customers = Customers.objects.filter(clientstatus='Active',
                                                         insuexpire__range=(from_date, to_date)).exclude(
                        insuexpire='0001-01-01').only('idcustomer', 'cusname', 'contact1', 'email', 'mobile1',
                                                      'insuexpire')
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': 'Ins',
                         'dateexp': customer.insuexpire.strftime('%m/%d/%Y')})
            if type_search == 'ANR':
                year = date.today().year
                customers = Customers.objects.filter(clientstatus='Active', anreport__lt=year).exclude(
                    Q(anreport='') | Q(anreport__isnull=True)).only('idcustomer', 'cusname', 'contact1', 'email',
                                                                    'mobile1', 'anreport').order_by('-anreport')
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': "Annual Report",
                         'dateexp': customer.anreport})
            if type_search == 'NM':
                if by == 'all':
                    customers = Customers.objects.filter(clientstatus='Active').exclude(
                        Q(mn='') | Q(mn__isnull=True)).only(
                        'idcustomer', 'cusname', 'contact1', 'email', 'mobile1', 'mn', 'mnexp')
                if by == 'date':
                    customers = Customers.objects.filter(clientstatus='Active',
                                                         mnexp__range=(from_date, to_date)).exclude(
                        Q(mn='') | Q(mn__isnull=True)).only(
                        'idcustomer', 'cusname', 'contact1', 'email', 'mobile1', 'mn', 'mnexp')
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.mn,
                         'dateexp': customer.mnexp.strftime('%m/%d/%Y')})
            if type_search == 'NY':
                customers = Customers.objects.filter(clientstatus='Active').exclude(
                    Q(nyid='') | Q(nyid__isnull=True)).only('idcustomer', 'cusname', 'contact1', 'email', 'mobile1',
                                                            'nyid', 'nyexp').order_by('nyexp')
                if by == 'date':
                    customers = customers.filter(nyexp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.nyid,
                         'dateexp': customer.nyexp.strftime('%m/%d/%Y')})
            if type_search == 'KYU':
                customers = Customers.objects.filter(clientstatus='Active').exclude(
                    Q(kyuid='') | Q(kyuid__isnull=True)).only('idcustomer', 'cusname', 'contact1', 'email', 'mobile1',
                                                              'kyuid', 'campcexp').order_by('campcexp')
                if by == 'date':
                    customers = customers.filter(campcexp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.kyuid,
                         'dateexp': customer.campcexp.strftime('%m/%d/%Y')})
            if type_search == 'UCR':
                customers = Customers.objects.filter(clientstatus='Active', mcexp__range=(from_date, to_date)).exclude(
                    Q(mc='') | Q(mc__isnull=True)).only(
                    'idcustomer', 'cusname', 'contact1', 'email', 'mobile1', 'mc', 'mcexp')
                if by == 'date':
                    customers = customers.filter(mcexp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.mc,
                         'dateexp': customer.mcexp.strftime('%m/%d/%Y')})
            if type_search == 'DOT':
                customers = Customers.objects.filter(clientstatus='Active').exclude(
                    Q(dotid='') | Q(dotid__isnull=True)).only('idcustomer', 'cusname', 'contact1', 'email', 'mobile1',
                                                              'dotid', 'dotidexp').order_by('dotidexp')
                if by == 'date':
                    customers = customers.filter(dotidexp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.dotid,
                         'dateexp': customer.dotidexp.strftime('%m/%d/%Y') if customer.dotidexp else customer.dotidexp})
            if type_search == 'RANDOM':
                customers = Customers.objects.filter(clientstatus='Active', dotclient='Yes').only('idcustomer',
                                                                                                  'cusname', 'contact1',
                                                                                                  'email', 'mobile1',
                                                                                                  'dotid',
                                                                                                  'bitexp').order_by(
                    'bitexp')
                if by == 'date':
                    customers = customers.filter(bitexp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {'idcustomer': customer.idcustomer, 'cusname': customer.cusname, 'contact1': customer.contact1,
                         'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.dotid,
                         'dateexp': customer.bitexp.strftime('%m/%d/%Y')})
            if type_search == 'INNER BRIDGE':
                customers = Customers.objects.filter(clientstatus='Active').exclude(
                    Q(inner_bridge_exp__isnull=True) | Q(inner_bridge_exp__year='0001')).only('idcustomer', 'cusname',
                                                                                              'contact1', 'email',
                                                                                              'mobile1',
                                                                                              'careg',
                                                                                              'inner_bridge_exp').order_by(
                    'inner_bridge_exp')
                if by == 'date':
                    customers = customers.filter(inner_bridge_exp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {
                            'idcustomer': customer.idcustomer, 'cusname': customer.cusname,
                            'contact1': customer.contact1,
                            'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.careg,
                            'dateexp': customer.inner_bridge_exp.strftime('%m/%d/%Y')
                        }
                    )
            if type_search == 'OVERWEIGHT':
                customers = Customers.objects.filter(clientstatus='Active') \
                    .exclude(Q(over_weight_exp__isnull=True) | Q(over_weight_exp__year='0001')) \
                    .only('idcustomer', 'cusname', 'contact1', 'email', 'mobile1',
                          'california', 'over_weight_exp').order_by('inner_bridge_exp')
                if by == 'date':
                    customers = customers.filter(over_weight_exp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {
                            'idcustomer': customer.idcustomer, 'cusname': customer.cusname,
                            'contact1': customer.contact1,
                            'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.california,
                            'dateexp': customer.over_weight_exp.strftime('%m/%d/%Y')
                        }
                    )
            if type_search == 'OVERSIZE':
                customers = Customers.objects.filter(clientstatus='Active') \
                    .exclude(Q(over_size_exp__isnull=True) | Q(over_size_exp__year='0001')) \
                    .only('idcustomer', 'cusname', 'contact1', 'email', 'mobile1',
                          'mcreg', 'over_size_exp').order_by('over_size_exp')
                if by == 'date':
                    customers = customers.filter(over_weight_exp__range=(from_date, to_date))
                for customer in customers:
                    data.append(
                        {
                            'idcustomer': customer.idcustomer, 'cusname': customer.cusname,
                            'contact1': customer.contact1,
                            'email': customer.email, 'mobile1': customer.mobile1, 'type': customer.mcreg,
                            'dateexp': customer.over_size_exp.strftime('%m/%d/%Y')
                        }
                    )

            return JsonResponse({'data': data}, safe=False)
    else:
        redirect("/Procedure/login/")


# Export to Excel
@login_required(login_url='Procedure:login')
def print_summary(request):
    if request.method == 'POST':
        idcustomer = request.POST.get('idcustomer')
        idunit = request.POST.get('idunit')
        year = request.POST.get('year')
        quarter = int(request.POST.get('quarter'))
        months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
                  'November', 'December']
        file_name = '{0}_{1}_{2}.xlsx'.format(idcustomer, idunit, date.today().strftime('%m%d%Y'))
        file_path = '{}/{}'.format(settings.FILES_PDF, file_name)
        if os.path.isfile(file_path):
            os.remove(file_path)
        first_month, second_month, third_month = receive_summary(idcustomer, idunit, year, quarter)
        book = Workbook()
        if quarter == 1:
            write_month = months[0]
        if quarter == 2:
            write_month = months[3]
        if quarter == 3:
            write_month = months[6]
        if quarter == 4:
            write_month = months[9]
        sheet = book.active
        sheet.title = write_month
        sheet.append(('State', 'Miles', 'Gallons'))
        for values in first_month:
            sheet.append((values['state'], values['mile'], values['gallon']))
        # Second Month
        if quarter == 1:
            write_month = months[1]
        if quarter == 2:
            write_month = months[4]
        if quarter == 3:
            write_month = months[7]
        if quarter == 4:
            write_month = months[10]
        sheet = book.create_sheet(write_month)
        sheet.append(('State', 'Miles', 'Gallons'))
        for values in second_month:
            sheet.append((values['state'], values['mile'], values['gallon']))
        # Third Month
        if quarter == 1:
            write_month = months[2]
        if quarter == 2:
            write_month = months[5]
        if quarter == 3:
            write_month = months[8]
        if quarter == 4:
            write_month = months[11]
        sheet = book.create_sheet(write_month)
        sheet.append(('State', 'Miles', 'Gallons'))
        for values in third_month:
            sheet.append((values['state'], values['mile'], values['gallon']))
        book.save(file_path)
        return JsonResponse({'filename': file_name})


class Sellers(View):
    content_type = ContentType.objects.get_for_model(Invoice_paid)
    Permission.objects.get_or_create(
        name="Sellers",
        codename="sellers",
        content_type=content_type,
    )
    template_name = 'Procedure/Reports/Sales/sellers.html'

    def get(self, request, *args, **kwargs):
        if  request.GET.get('action') == 'details':
            action = request.GET.get('action')
            if action == 'details':
                user_id = request.GET.get('user_id')
                html = render_to_string('Procedure/Reports/Sales/sales.html',
                                        {'user_id': user_id, 'from_date': request.GET.get('from_date'),
                                         'to_date': request.GET.get('to_date')})
                return JsonResponse({
                    'html': html,
                    'parent-row-id': user_id
                })
        else:
            return render(request, self.template_name, {})

    def post(self, request, *args, **kwargs):
        data = list()
        num = 0
        if request.POST :

            if request.POST.get('from_date') and request.POST.get('to_date'):
                start_date = datetime.strptime(request.POST.get('from_date'), '%m/%d/%Y').date()
                end_date = datetime.strptime(request.POST.get('to_date'), '%m/%d/%Y').date()
            else:
                today = datetime.today()
                start_date = datetime.date(today.year, today.month, 1)
                end_date = today.strftime('%Y-%m-%d')
            query = "SELECT i.idinvoice, i.iduser AS iduser, COUNT(i.iduser) AS numinvoice, SUM(i.amount) AS amount, u.fullname AS fullname, u.avatar AS photo, u.is_active FROM invoices i INNER JOIN user u ON iduser=u.id WHERE ((i.invdate BETWEEN '{0}' AND '{1}') OR (i.paid_date BETWEEN '{0}' AND '{1}')) AND (i.deu = 0 AND i.status='Paid' AND i.deleted = False) GROUP BY i.iduser ORDER BY i.iduser ASC;".format(
                start_date, end_date)
            sales = Invoices.objects.raw(query)
            data = list()
            for sale in sales:
                data.append({
                    'Id': sale.iduser_id, 'Name': sale.fullname.upper(), 'Sales': sale.numinvoice, 'Photo': sale.photo,
                    'Amount': round(sale.amount, 2),
                    'FromDate': start_date, 'ToDate': end_date, 'is_active': sale.is_active
                })
            num = len(sales)
            return JsonResponse({'draw': 1, 'recordsTotal': num, 'recordsFiltered': num, 'data': data}, safe=False)
        else:
            return JsonResponse({'draw': 1, 'recordsTotal': num, 'recordsFiltered': num, 'data': data}, safe=False)
